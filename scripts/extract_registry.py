#!/usr/bin/env python3
"""Extract the CRMDR canonical ID registry from the private digitization workbook.

The workbook ("Roman Martyrology LA IT EN with IDs.xlsx") contains the full
Latin, Italian and English texts of the eulogies of the Roman Martyrology
(editio altera 2004). Those texts are copyrighted and are NOT extracted here:
this script reads only the non-copyrightable structural metadata — the
canonical ID, calendar placement (month/day/entry), the asterisk marker, and
the country associated with the place of the elogium — and writes:

  data/martyrology_ids.json   machine-readable registry
  registry/MM-<month>.md      human-readable per-month tables

Usage:
  python3 extract_registry.py /path/to/"Roman Martyrology LA IT EN with IDs.xlsx" [repo_root]

Requires: openpyxl
"""

import json
import sys
from pathlib import Path

import openpyxl

MONTH_SHEETS = [
    "GENNAIO", "FEBBRAIO", "MARZO", "APRILE", "MAGGIO", "GIUGNO",
    "LUGLIO", "AGOSTO", "SETTEMBRE", "OTTOBRE", "NOVEMBRE", "DICEMBRE",
]
MONTH_NAMES_EN = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# IDs in the workbook that violate the canonicalization rules and are
# corrected on extraction. Surnames are not latinized unless an already
# well-known Latin form exists: "Miki" is a Japanese surname with no such form.
ID_CORRECTIONS = {
    "mr:0206-paulus-mikus-et-socii": "mr:0206-paulus-miki-et-socii",
    # The Latin editio altera 2004 places Bl. Marcantonio Durando at June 10
    # (entry 9*); the Italian (CEI) edition places him at December 10 (entry
    # 9*). The MMDD anchor follows the Latin print.
    "mr:1210-marcus-antonius-durando": "mr:0610-marcus-antonius-durando",
    # The elogium's first-named subject is San Guido, abbot; "Domninus" comes
    # from the place name (Burgi Sancti Domnini, Borgo San Donnino, today
    # Fidenza). Rule 1: the slug is the first-named subject.
    "mr:0331-domninus": "mr:0331-guido",
    # Polish ł/Ł (U+0142/U+0141) is a precomposed letter, not base l + a
    # combining stroke, so NFKD normalization leaves it intact and the
    # diacritic-folding step that produced the workbook slugs dropped it to a
    # blank, splitting the surname (paw-owski). The standard ASCII
    # transliteration is a plain l (Pawłowski -> pawlowski). The folding logic
    # is fixed in extract_subjects.py; these entries correct the slugs the
    # buggy fold already baked into the workbook.
    "mr:0109-iosephus-paw-owski-et-casimirus-grelewskus": "mr:0109-iosephus-pawlowski-et-casimirus-grelewskus",
    "mr:0219-iosephus-zap-ata": "mr:0219-iosephus-zaplata",
    "mr:0308-vincentius-kad-ubek": "mr:0308-vincentius-kadlubek",
    "mr:0331-natalia-tu-asiewicz": "mr:0331-natalia-tulasiewicz",
    "mr:0719-achilles-pucha-a-et-hermannus-stepien": "mr:0719-achilles-puchala-et-hermannus-stepien",
    "mr:0731-michael-ozieb-owski": "mr:0731-michael-ozieblowski",
    "mr:0908-ladislaus-b-adzinski": "mr:0908-ladislaus-bladzinski",
    "mr:1014-stanislaus-mysakowski-et-franciscus-ros-aniec": "mr:1014-stanislaus-mysakowski-et-franciscus-roslaniec",
    "mr:1216-honoratus-de-bia-a-podlaska-kazminsky": "mr:1216-honoratus-de-biala-podlaska-kazminsky",
    "mr:1219-maria-eva-de-providentia-noiszewska-et-maria-martha-de-iesu-wo-owsk": "mr:1219-maria-eva-de-providentia-noiszewska-et-maria-martha-de-iesu-wolowsk",
    # Short romanized surnames were latinized in the workbook slugs by stripping
    # the final vowel and appending -us (Di -> dus, Li -> lus, Yi -> yus,
    # Xi -> xus, Bùi -> buus, Miki -> mikus). Rule: surnames are not latinized,
    # so the romanized form is preserved as printed in the elogium. (The Feb 6
    # Paul Miki memorial is already corrected above; mr:0205 is the Feb 5 pridie
    # anticipation the earlier correction missed.)
    "mr:0109-agatha-yus": "mr:0109-agatha-yi",
    "mr:0110-aegidius-dus-bello": "mr:0110-aegidius-di-bello",
    "mr:0121-ioannes-yus-yun-il": "mr:0121-ioannes-yi-yun-il",
    "mr:0205-paulus-mikus-et-socii": "mr:0205-paulus-miki-et-socii",
    "mr:0219-lucia-yus-zhenmei": "mr:0219-lucia-yi-zhenmei",
    "mr:0524-augustinus-yus-kwang-hon": "mr:0524-augustinus-yi-kwang-hon",
    "mr:0601-hannibalis-maria-dus-francia": "mr:0601-hannibalis-maria-di-francia",
    "mr:0613-augustinus-phan-viet-huy-et-nicolaus-buus-viet-the": "mr:0613-augustinus-phan-viet-huy-et-nicolaus-bui-viet-the",
    "mr:0630-raymundus-lus-quanzhen-et-petrus-lus-quanhui": "mr:0630-raymundus-li-quanzhen-et-petrus-li-quanhui",
    "mr:0720-magdalena-yus-yong-hui-et-socii": "mr:0720-magdalena-yi-yong-hui-et-socii",
    "mr:0720-xus-guizi": "mr:0720-xi-guizi",
    "mr:1125-petrus-yus-ho-yong": "mr:1125-petrus-yi-ho-yong",
    "mr:1219-franciscus-xaverius-ha-trong-mau-et-dominicus-buus-van-uy": "mr:1219-franciscus-xaverius-ha-trong-mau-et-dominicus-bui-van-uy",
    # Byname disambiguation: Gregory of Nyssa (well known by that byname)
    # shares Jan 10 with Pope Gregory X (mr:0110-gregorius-x). Rule 1's bare
    # first-named subject "gregorius" is ambiguous here, so the byname is kept.
    "mr:0110-gregorius": "mr:0110-gregorius-nyssenus",
}

# Days whose opening elogia are printed as unnumbered drop-cap paragraphs in
# both editions (celebrations with liturgical rank: solemnities, feasts,
# memorias — including days with two or three memorias). Maps (month, day) ->
# number of unnumbered leading entries; the printed numbering counts them
# implicitly, so the workbook entry numbers stay aligned. Derived mechanically
# from the full-corpus sweep of both OCR layers (July 2026).
UNNUMBERED_LEADS = {
    (1, 1): 1, (1, 2): 1, (1, 3): 1, (1, 6): 1, (1, 7): 1, (1, 13): 1,
    (1, 17): 1, (1, 20): 2, (1, 21): 1, (1, 22): 1, (1, 24): 1, (1, 25): 1,
    (1, 26): 1, (1, 27): 1, (1, 28): 1, (1, 31): 1, (2, 2): 1, (2, 3): 2,
    (2, 5): 1, (2, 6): 1, (2, 8): 2, (2, 10): 1, (2, 11): 1, (2, 14): 1,
    (2, 17): 1, (2, 21): 1, (2, 22): 1, (2, 23): 1, (3, 4): 1, (3, 7): 1,
    (3, 8): 1, (3, 9): 1, (3, 17): 1, (3, 18): 1, (3, 19): 1, (3, 23): 1,
    (3, 25): 1, (4, 2): 1, (4, 4): 1, (4, 5): 1, (4, 7): 1, (4, 11): 1,
    (4, 13): 1, (4, 21): 1, (4, 23): 2, (4, 24): 1, (4, 25): 1, (4, 28): 2,
    (4, 29): 1, (4, 30): 1, (5, 1): 1, (5, 2): 1, (5, 3): 1, (5, 12): 2,
    (5, 13): 1, (5, 14): 1, (5, 18): 1, (5, 20): 1, (5, 21): 1, (5, 22): 1,
    (5, 25): 3, (5, 26): 1, (5, 27): 1, (5, 31): 1, (6, 1): 1, (6, 2): 1,
    (6, 3): 1, (6, 5): 1, (6, 6): 1, (6, 9): 1, (6, 11): 1, (6, 13): 1,
    (6, 19): 1, (6, 21): 1, (6, 22): 2, (6, 24): 1, (6, 27): 1, (6, 28): 1,
    (6, 29): 1, (6, 30): 1, (7, 3): 1, (7, 4): 1, (7, 5): 1, (7, 6): 1,
    (7, 9): 1, (7, 11): 1, (7, 13): 1, (7, 14): 1, (7, 15): 1, (7, 16): 1,
    (7, 20): 1, (7, 21): 1, (7, 22): 1, (7, 23): 1, (7, 24): 1, (7, 25): 1,
    (7, 26): 1, (7, 29): 1, (7, 30): 1, (7, 31): 1, (8, 1): 1, (8, 2): 2,
    (8, 4): 1, (8, 5): 1, (8, 6): 1, (8, 7): 2, (8, 8): 1, (8, 9): 1,
    (8, 10): 1, (8, 11): 1, (8, 12): 1, (8, 13): 1, (8, 14): 1, (8, 15): 1,
    (8, 16): 1, (8, 19): 1, (8, 20): 1, (8, 21): 1, (8, 22): 1, (8, 23): 1,
    (8, 24): 1, (8, 25): 2, (8, 27): 1, (8, 28): 1, (8, 29): 1, (9, 3): 1,
    (9, 8): 1, (9, 9): 1, (9, 12): 1, (9, 13): 1, (9, 14): 1, (9, 15): 1,
    (9, 16): 1, (9, 17): 1, (9, 19): 1, (9, 20): 1, (9, 21): 1, (9, 23): 1,
    (9, 26): 1, (9, 27): 1, (9, 28): 2, (9, 29): 1, (9, 30): 1, (10, 1): 1,
    (10, 2): 1, (10, 4): 1, (10, 6): 1, (10, 7): 1, (10, 9): 2, (10, 14): 1,
    (10, 15): 1, (10, 16): 2, (10, 17): 1, (10, 18): 1, (10, 19): 2, (10, 23): 1,
    (10, 24): 1, (10, 28): 1, (11, 1): 1, (11, 2): 1, (11, 3): 1, (11, 4): 1,
    (11, 9): 1, (11, 10): 1, (11, 11): 1, (11, 12): 1, (11, 15): 1, (11, 16): 2,
    (11, 17): 1, (11, 18): 1, (11, 21): 1, (11, 22): 1, (11, 23): 2, (11, 24): 1,
    (11, 25): 1, (11, 30): 1, (12, 3): 1, (12, 4): 1, (12, 6): 1, (12, 7): 1,
    (12, 8): 1, (12, 9): 1, (12, 11): 1, (12, 12): 1, (12, 13): 1, (12, 14): 1,
    (12, 21): 1, (12, 23): 1, (12, 25): 1, (12, 26): 1, (12, 27): 1, (12, 28): 1,
    (12, 29): 1, (12, 31): 1,
}

# Placement overrides: the digitized workbook records the Italian (CEI)
# placement, but the anchor edition (the Latin editio altera 2004 print)
# places the elogium on a different day. Maps corrected ID -> (month, day,
# entry, note). `entry` is None because the workbook has no row for the
# anchor-edition day.
PLACEMENT_OVERRIDES = {
    "mr:0610-marcus-antonius-durando": (
        6, 10, None,
        "Entry 9* at June 10 in the Latin editio altera 2004 print (verified "
        "on the page scan); the Italian (CEI) edition and the digitized "
        "workbook place the elogium at December 10 (entry 9*).",
    ),
}

# Entries present in the Italian (CEI) edition and the digitized workbook but
# absent from the Latin editio altera 2004 print (all verified on the page
# scans of both editions, July 2026).
ENTRY_NOTES = {
    "mr:0712-proclus-et-hilarion": (
        "Entry 1 at July 12 in the Italian (CEI) edition; absent from the "
        "Latin editio altera 2004 print, whose July 12 numbering begins at 2 "
        "with the gap left unrenumbered."
    ),
    "mr:0825-eusebius-et-socii": (
        "Entry 3 at August 25 in the Italian (CEI) edition; absent from the "
        "Latin editio altera 2004 print, where the day's numbered entries "
        "begin at 3 (Genesius) after the two unnumbered memorias."
    ),
    "mr:0709-maria-a-iesu-crucifixo-petkovic": (
        "Entry 11* at July 9 in the Italian (CEI) edition; absent from the "
        "Latin editio altera 2004 print, whose July 9 ends at entry 10*. "
        "Bl. Marija Petković was beatified on 6 June 2003."
    ),
}

# Asterisk overrides where the digitized workbook follows the Italian (CEI)
# edition but the anchor edition (the Latin editio altera 2004 print) differs.
# All were verified against the page scans of BOTH printed editions (2026-07),
# independently of the workbook: asterisk-presence claims via the OCR text
# layer with visual spot-checks, asterisk-absence claims each visually
# confirmed on the scan of the edition concerned.

# Entries asterisked in the Latin print but not in the CEI edition (nor in the
# workbook). Maps ID -> entry number as printed in the Latin editio altera.
LATIN_ASTERISKED = {
    "mr:0104-ferreolus": 4,
    "mr:0104-rigomerus": 5,
    "mr:0104-pharaildis": 7,
    "mr:0122-ladislaus-batthyany-strattmann": 15,
    "mr:0502-boleslaus-strzelecki": 12,
    "mr:0512-imelda-lambertini": 10,
    "mr:0524-servulus": 4,
    "mr:0611-bardo": 4,
    "mr:0624-gohardus": 7,
    "mr:0711-leontius": 5,
    "mr:0718-tarsicia-mackiv": 13,
    "mr:0730-godeleva": 6,
    "mr:0802-betharius": 8,
    "mr:0807-donatus-vesontione": 7,
    "mr:0909-franciscus-garate-aranguren": 10,
    "mr:1021-wendelinus": 8,
    "mr:1026-eata": 7,
    "mr:1103-libertinus": 3,
    "mr:1103-odrada": 10,
    "mr:1118-theofredus": 6,
    "mr:1119-eudo": 6,
    "mr:1215-marinus": 3,
    "mr:1230-egwinus": 7,
}

# Entries with a plain number in the Latin print that the CEI edition (and the
# workbook) marks with an asterisk. Maps ID -> entry number in the Latin print.
LATIN_PLAIN = {
    "mr:0323-rebecca-de-himlaya": 11,
    "mr:0812-iacobus": 11,
    "mr:0821-iosephus": 11,
    "mr:1201-domnolus": 5,
    "mr:1203-lucius": 5,
    "mr:1212-simon-phan": 11,
}

ASTERISK_OVERRIDES = {}
for _id, _n in LATIN_ASTERISKED.items():
    ASTERISK_OVERRIDES[_id] = (
        True,
        f"Asterisked entry ({_n}*) in the Latin editio altera 2004 print; "
        "the Italian (CEI) edition carries no asterisk.",
    )
for _id, _n in LATIN_PLAIN.items():
    ASTERISK_OVERRIDES[_id] = (
        False,
        f"Plain entry ({_n}., no asterisk) in the Latin editio altera 2004 "
        "print, visually verified on the page scan; the Italian (CEI) edition "
        "marks the entry with an asterisk.",
    )

# Entries present in the editio altera 2004 print but absent from the
# digitized "with IDs" workbook (see docs/canonicalization-report.md). Both
# are present in the parallel-texts workbook, whose reviewer comments document
# their numbering across editions.
PRINT_ONLY_ENTRIES = [
    {
        "id": "mr:0104-abrunculus",
        "month": 1,
        "day": 4,
        "entry": None,
        "asterisk": True,
        "country": "FR",
        "note": "Entry 2* in the Latin editio altera 2004 print; absent from "
                "the Italian (CEI) edition, from Mons. Barba's Word "
                "transcription, and from the digitized workbook.",
    },
    {
        "id": "mr:0104-emmanuel-gonzalez-garcia",
        "month": 1,
        "day": 4,
        "entry": None,
        "asterisk": True,
        "country": "ES",
        "note": "Numbered 12* in the Latin editio altera 2004 print, 11* in "
                "the Italian (CEI) edition and in Mons. Barba's Word "
                "transcription; absent from the digitized workbook. "
                "Bl. Manuel González García was canonized in 2016: status "
                "change with no ID change.",
    },
]

# The four leap-day elogia are printed twice (Feb 28 and Feb 29) and carry a
# single identity each, anchored at 0229. The Feb 29 placement is primary;
# the Feb 28 placement is recorded in `also_on`.
LEAP_DAY_PRIMARY = (2, 29)


def build_country_map(wb):
    iso = {}
    ws = wb["Paesi"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, _en, it = (list(row) + [None] * 3)[:3]
        if code and it:
            iso[it.strip()] = code.strip()
    return iso


def extract(workbook_path):
    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    iso = build_country_map(wb)
    rows = []
    for month_index, sheet in enumerate(MONTH_SHEETS, 1):
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=3, values_only=True):
            _mese, giorno, voce, asterisk, _it, _en, _la, paese, mr_id = (
                list(row) + [None] * 9
            )[:9]
            if mr_id is None:
                continue
            paese = str(paese).strip() if paese is not None else ""
            country = iso[paese] if paese else None
            mr_id = str(mr_id).strip()
            mr_id = ID_CORRECTIONS.get(mr_id, mr_id)
            row_out = {
                "id": mr_id,
                "month": month_index,
                "day": int(str(giorno)),
                "entry": int(str(voce)),
                "asterisk": asterisk == "*",
                "country": country,
            }
            if mr_id in ASTERISK_OVERRIDES:
                row_out["asterisk"], row_out["note"] = ASTERISK_OVERRIDES[mr_id]
            if mr_id in PLACEMENT_OVERRIDES:
                p_month, p_day, p_entry, p_note = PLACEMENT_OVERRIDES[mr_id]
                row_out.update(month=p_month, day=p_day, entry=p_entry)
                row_out["note"] = (row_out.get("note", "") + " " + p_note).strip()
            if mr_id in ENTRY_NOTES:
                row_out["note"] = (row_out.get("note", "") + " " + ENTRY_NOTES[mr_id]).strip()
            leads = UNNUMBERED_LEADS.get((row_out["month"], row_out["day"]), 0)
            if row_out["entry"] is not None and row_out["entry"] <= leads:
                row_out["unnumbered"] = True
            rows.append(row_out)

    # Merge duplicate IDs (the leap-day elogia): keep the Feb 29 placement as
    # primary and record the other placement in `also_on`.
    by_id = {}
    entries = []
    for row in rows:
        if row["id"] in by_id:
            first = by_id[row["id"]]
            primary, secondary = (
                (row, first)
                if (row["month"], row["day"]) == LEAP_DAY_PRIMARY
                else (first, row)
            )
            primary["also_on"] = [{
                "month": secondary["month"],
                "day": secondary["day"],
                "entry": secondary["entry"],
            }]
            if first is not primary:
                entries[entries.index(first)] = primary
            by_id[row["id"]] = primary
        else:
            by_id[row["id"]] = row
            entries.append(row)

    entries.extend(PRINT_ONLY_ENTRIES)
    entries.sort(key=lambda e: (e["month"], e["day"], e["entry"] is None, e["entry"] or 0))
    return entries


def load_deprecated(repo_root, current_ids):
    """Deprecated canonical IDs coined for eulogies of historical editions
    with no counterpart in the anchor edition (data/deprecated_ids.json,
    generated by the martyrology-api alignment tooling). Their MMDD anchors
    the placement in the edition named by attested_in."""
    path = repo_root / "data" / "deprecated_ids.json"
    if not path.exists():
        return []
    deprecated = json.load(open(path, encoding="utf-8"))
    for e in deprecated:
        assert e.get("deprecated") is True and e["id"] not in current_ids, e["id"]
    return deprecated


def write_json(entries, deprecated, repo_root):
    out = {
        "$comment": "Canonical ID registry for the eulogies of the Roman Martyrology. "
                    "IDs are drafts pending committee review; see the README and "
                    "docs/canonicalization-report.md. Entries with deprecated:true "
                    "are attested only in historical editions (their MMDD anchors "
                    "the placement in the edition named by attested_in); the "
                    "deprecated status is itself a mechanical draft.",
        "anchor_edition": "martyrologium_romanum_editio_altera_2004",
        "id_scheme": "mr:MMDD-slug",
        "entry_count": len(entries) + len(deprecated),
        "current_count": len(entries),
        "deprecated_count": len(deprecated),
        "entries": entries + deprecated,
    }
    path = repo_root / "data" / "martyrology_ids.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
        f.write("\n")
    return path


def write_markdown(entries, repo_root):
    reg_dir = repo_root / "registry"
    reg_dir.mkdir(parents=True, exist_ok=True)
    for month_index, month_name in enumerate(MONTH_NAMES_EN, 1):
        month_entries = [e for e in entries if e["month"] == month_index]
        lines = [
            f"# {month_name}",
            "",
            f"{len(month_entries)} canonical IDs. "
            "`Entry` is the elogium's position within the day in the digitized workbook "
            "(editio altera 2004); an entry number in parentheses marks an unnumbered "
            "header elogium (a drop-cap paragraph for a celebration with liturgical "
            "rank, counted but not printed as a number); `*` marks asterisked entries; "
            "`Country` is the ISO 3166-1 alpha-2 code of the modern country of the "
            "place of the elogium.",
            "",
            "| Day | Entry | ID | * | Country | Notes |",
            "| --- | --- | --- | --- | --- | --- |",
        ]
        for e in month_entries:
            notes = []
            if e.get("also_on"):
                for a in e["also_on"]:
                    notes.append(f"also printed at {a['month']}/{a['day']} entry {a['entry']}")
            if e.get("note"):
                notes.append(e["note"])
            if e["entry"] is None:
                entry_cell = "—"
            elif e.get("unnumbered"):
                entry_cell = f"({e['entry']})"
            else:
                entry_cell = str(e["entry"])
            lines.append(
                f"| {e['day']} | {entry_cell} "
                f"| `{e['id']}` | {'*' if e['asterisk'] else ''} "
                f"| {e['country'] or ''} | {' '.join(notes)} |"
            )
        path = reg_dir / f"{month_index:02d}-{month_name.lower()}.md"
        path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    workbook_path = sys.argv[1]
    repo_root = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(__file__).resolve().parent.parent
    entries = extract(workbook_path)
    deprecated = load_deprecated(repo_root, {e["id"] for e in entries})
    path = write_json(entries, deprecated, repo_root)
    write_markdown(entries, repo_root)
    ids = [e["id"] for e in entries] + [e["id"] for e in deprecated]
    assert len(ids) == len(set(ids)), "duplicate IDs after merge"
    print(f"Wrote {len(entries)} current + {len(deprecated)} deprecated entries "
          f"to {path} and registry/*.md")


if __name__ == "__main__":
    main()
